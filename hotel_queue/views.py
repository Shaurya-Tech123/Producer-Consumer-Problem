from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.http import HttpResponse
from django.conf import settings
from django.contrib.auth.models import User
from django.contrib.auth import authenticate
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

from .forms import DemandForm
from .models import Demand, HotelSettings, StaffRole, StaffMember


@login_required
def dashboard(request):
    # Only show pending and in-progress demands in the dashboard
    demands = Demand.objects.exclude(status="Completed")
    form = DemandForm()
    settings_obj = HotelSettings.objects.first()
    return render(
        request,
        "hotel_queue/dashboard.html",
        {"demands": demands, "form": form, "settings": settings_obj},
    )


@login_required
def add_demand(request):
    if not request.user.is_staff:
        messages.error(request, "Only staff can add demands.")
        return redirect("dashboard")
    if request.method == "POST":
        form = DemandForm(request.POST)
        if form.is_valid():
            demand = form.save(commit=False)
            demand.created_by = request.user
            demand.save()
            messages.success(request, "Demand added.")
            return redirect("dashboard")
        # Re-render dashboard with errors, only showing pending and in-progress demands
        demands = Demand.objects.exclude(status="Completed")
        messages.error(request, "Please correct the errors below.")
        # We no longer need to set different choices based on demand type
        # as we now use a combined dropdown for all locations
        return render(request, "hotel_queue/dashboard.html", {"demands": demands, "form": form})
    return redirect("dashboard")


@login_required
def settings_page(request):
    if not request.user.is_staff:
        messages.error(request, "Settings are for staff only.")
        return redirect("dashboard")
    # Simple settings editor: change table/room counts and list staff/roles
    settings_obj, _ = HotelSettings.objects.get_or_create(id=1)
    roles = StaffRole.objects.all()
    staff = StaffMember.objects.select_related("role").all()
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "update_counts":
            num_tables = int(request.POST.get("num_tables", settings_obj.num_tables) or settings_obj.num_tables)
            num_rooms = int(request.POST.get("num_rooms", settings_obj.num_rooms) or settings_obj.num_rooms)
            settings_obj.num_tables = num_tables
            settings_obj.num_rooms = num_rooms
            settings_obj.save()
            messages.success(request, "Counts updated.")
        elif action == "add_role":
            role_name = (request.POST.get("role_name") or "").strip()
            if role_name:
                StaffRole.objects.get_or_create(name=role_name)
                messages.success(request, "Role added.")
            else:
                messages.error(request, "Role name required.")
        elif action == "delete_role":
            role_id = request.POST.get("role_id")
            if role_id:
                StaffRole.objects.filter(id=role_id).delete()
                messages.success(request, "Role deleted.")
        elif action == "add_staff":
            staff_name = (request.POST.get("staff_name") or "").strip()
            role_id = request.POST.get("staff_role_id")
            if staff_name and role_id:
                try:
                    role = StaffRole.objects.get(id=role_id)
                    StaffMember.objects.create(name=staff_name, role=role)
                    messages.success(request, "Staff added.")
                except StaffRole.DoesNotExist:
                    messages.error(request, "Invalid role.")
            else:
                messages.error(request, "Staff name and role required.")
        elif action == "delete_staff":
            staff_id = request.POST.get("staff_id")
            if staff_id:
                StaffMember.objects.filter(id=staff_id).delete()
                messages.success(request, "Staff deleted.")
        return redirect("settings_page")
    roles = StaffRole.objects.all()
    staff = StaffMember.objects.select_related("role").all()
    return render(request, "hotel_queue/settings.html", {"settings": settings_obj, "roles": roles, "staff": staff})


@login_required
def mark_in_progress(request, pk):
    if not request.user.is_staff:
        messages.error(request, "Only staff can update demand status.")
        return redirect("dashboard")
    demand = get_object_or_404(Demand, pk=pk)
    demand.status = "In Progress"
    demand.fulfilled_by = request.user
    demand.save(update_fields=["status", "fulfilled_by"])
    messages.info(request, "Marked as In Progress.")
    return redirect("dashboard")


@login_required
def mark_completed(request, pk):
    if not request.user.is_staff:
        messages.error(request, "Only staff can update demand status.")
        return redirect("dashboard")
    demand = get_object_or_404(Demand, pk=pk)
    demand.status = "Completed"
    demand.fulfilled_by = request.user
    demand.completed_at = timezone.now()
    demand.save(update_fields=["status", "fulfilled_by", "completed_at"])
    messages.success(request, "Marked as Completed.")
    return redirect("dashboard")


@login_required
def completed_list(request):
    demands = Demand.objects.filter(status="Completed").order_by("-completed_at")
    return render(request, "hotel_queue/completed.html", {"demands": demands})


@login_required
def export_completed_to_excel(request):
    if not request.user.is_staff:
        messages.error(request, "Only staff can export data.")
        return redirect("completed_list")
    
    # Get completed demands
    demands = Demand.objects.filter(status="Completed").order_by("-completed_at")
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Completed Tasks"
    
    # Define headers
    headers = [
        "ID", "Type", "Description", "Location", "Created At", 
        "Completed At", "Time Taken", "Created By", "Fulfilled By"
    ]
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add data
    for row, demand in enumerate(demands, 2):
        time_taken = ""
        if demand.completed_at:
            time_taken = str(demand.completed_at - demand.created_at).split('.')[0]
        
        ws.cell(row=row, column=1, value=demand.id)
        ws.cell(row=row, column=2, value=demand.demand_type)
        ws.cell(row=row, column=3, value=demand.description)
        ws.cell(row=row, column=4, value=demand.room_or_table or "-")
        ws.cell(row=row, column=5, value=demand.created_at.strftime("%Y-%m-%d %H:%M"))
        ws.cell(row=row, column=6, value=demand.completed_at.strftime("%Y-%m-%d %H:%M") if demand.completed_at else "-")
        ws.cell(row=row, column=7, value=time_taken)
        ws.cell(row=row, column=8, value=demand.created_by.username)
        ws.cell(row=row, column=9, value=demand.fulfilled_by.username if demand.fulfilled_by else "-")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"completed_tasks_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    return response


@login_required
def clear_completed_tasks(request):
    if not request.user.is_staff:
        messages.error(request, "Only staff can clear data.")
        return redirect("completed_list")
    
    if request.method == "POST":
        # Count tasks before deletion
        count = Demand.objects.filter(status="Completed").count()
        
        # Delete completed tasks
        Demand.objects.filter(status="Completed").delete()
        
        messages.success(request, f"Successfully cleared {count} completed tasks.")
        return redirect("completed_list")
    
    # If GET request, show confirmation
    count = Demand.objects.filter(status="Completed").count()
    return render(request, "hotel_queue/clear_confirm.html", {"count": count})



# TEMPORARY: Helper to reset/create viewer credentials in DEBUG only
def reset_viewer_credentials(request):
    if not settings.DEBUG:
        return HttpResponse("Disabled", status=403)
    viewer, created = User.objects.get_or_create(username="viewer", defaults={"is_staff": False, "is_active": True})
    viewer.is_staff = False
    viewer.is_active = True
    viewer.set_password("viewer12345")
    viewer.save()
    msg = "created" if created else "updated"
    return HttpResponse(f"viewer {msg} with password viewer12345", content_type="text/plain")

def debug_whoami(request):
    if not settings.DEBUG:
        return HttpResponse("Disabled", status=403)
    if request.user.is_authenticated:
        return HttpResponse(f"authenticated as: {request.user.username}")
    return HttpResponse("anonymous")

def debug_check_viewer(request):
    if not settings.DEBUG:
        return HttpResponse("Disabled", status=403)
    report = []
    try:
        viewer = User.objects.get(username="viewer")
        report.append(f"viewer exists; is_active={viewer.is_active}; is_staff={viewer.is_staff}")
    except User.DoesNotExist:
        return HttpResponse("viewer does not exist", status=404)
    pwd = request.GET.get("pwd", "viewer12345")
    user = authenticate(request, username="viewer", password=pwd)
    if user is not None:
        report.append("authenticate(viewer, provided pwd) -> OK")
    else:
        report.append("authenticate(viewer, provided pwd) -> FAILED")
    return HttpResponse("\n".join(report), content_type="text/plain")

